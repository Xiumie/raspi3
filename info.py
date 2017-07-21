#!/usr/bin/env python
# -*- coding: utf-8 -*-
# coding = utf-8

import urllib2
import urllib
import cookielib
import sys
import time
import os
import openpyxl
import random
from PIL import Image
import pytesseract

class SDU(object):
    def __init__(self):
        # 提交账号地址
        self.posturl = 'XXXXXXXXXXXXXXX'
        # 验证码地址
        self.captchaurl = 'XXXXXXXXXXXXXXX'
        # 学籍信息地址
        self.infourl = 'XXXXXXXXXXXXXXX'
        # 个人照片地址
        self.imgurl = 'XXXXXXXXXXXXXXX/xjInfoAction.do?oper=img'
        # 个人全部成绩地址
        self.gradeurl = 'XXXXXXXXXXXXXXX/gradeLnAllAction.do?type=ln&oper=fainfo&fajhh=403'
        # 根据抓包信息 构造headers
        self.headers = {
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.8,en;q=0.6',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded',
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.106 Safari/537.36',
        }

    def run(self, username, password):
        self.data_path = './data/' + username + '/'
        try:
            if os.path.exists(self.data_path):
                pass
            else:
                os.makedirs(self.data_path)
            # print os.getcwd()
            # 用户名和密码
            self.username = username
            self.password = password
            # 将cookies绑定到一个opener  cookie由cookielib自动管理
            self.cookie = cookielib.CookieJar()
            self.handler = urllib2.HTTPCookieProcessor(self.cookie)
            self.opener = urllib2.build_opener(self.handler)

            # 用openr访问验证码地址,获取cookie
            picture = self.opener.open(self.captchaurl).read()
            # 保存验证码到本地
            self.local = open('./temp/image.jpg', 'wb')
            self.local.write(picture)
            self.local.close()
            time.sleep(0.5)
            # 打开保存的验证码图片 输入
            self.SecretCode = pytesseract.image_to_string(Image.open('./temp/image.jpg'), lang='fra')
            # print self.SecretCode

            self.postdata = urllib.urlencode({
                'zjh': self.username,
                'mm': self.password,
                'v_yzm': self.SecretCode
            })
            # 获取文件
            info_grade = self.data_path + self.username + '_grade.html'
            info_info = self.data_path + self.username + '_info.html'
            info_img = self.data_path + self.username + '.jpg'
            # 构造request请求
            request = urllib2.Request(
                self.posturl,
                self.postdata,
                self.headers)
            # 向登录页面提交信息
            response = self.opener.open(request)
            # 下载学籍信息，全部成绩，个人照片
            grade_data = self.opener.open(self.gradeurl).read()
            with open(info_grade, 'w') as fout:
                fout.write(grade_data.encode('gbk'))
            info_data = self.opener.open(self.infourl).read()
            with open(info_info, 'w') as fout:
                fout.write(info_data.encode('gbk'))
            img_data = self.opener.open(self.imgurl).read()
            with open(info_img, 'wb') as fout:
                fout.write(img_data)
            return True
        except :
            # os.rmdir(data_path)
            return False

    def getxl_run(self):
        cnt = 0
        flg_ck = 0
        excel_path = 'XXXXXXXXXXXXXXX.xlsx'
        wb = openpyxl.load_workbook(excel_path)
        ck_data = wb.active
        sheet = wb.get_sheet_by_name('Sheet1')
        for row in range(2, sheet.max_row+1):
            cnt = 0
            flg_ck = 0
            li = sheet.cell(row=row, column=3).value
            if li is not None:
                while cnt < 20:
                    if sdu.run(li, li) is False:
                        cnt = cnt + 1
                        flg_ck = 0
                        # print cnt
                        time.sleep(random.randint(1, 10))
                    else:
                        flg_ck = 1
                        break 
                if flg_ck == 1:
                    print li + ' done'
                    ck_data.cell(row=row, column=11).value = 1
                else:
                    try:
                        os.rmdir(self.data_path)
                    except:
                        pass
                    ck_data.cell(row=row, column=11).value = 0
                # print sheet.cell(row=row, column=3).value
            else:
                pass
        wb.save('XXXXXXXXXXXXXXX.xlsx')

if __name__ == '__main__':
    reload(sys)
    sys.setdefaultencoding("gbk")
    sdu = SDU()
    sdu.getxl_run()
